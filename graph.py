import os
import re
import openpyxl as xl
import pandas as pd


class Graph:
    """
    running 'wikiscraper.main' yields a [root node].txt file for each sub-tree.
    So, each tree is a directory. Each sub-tree is a file within that directory.
    The urls contained within each [root node] file are branches (which may themselves be
    roots to other sub-trees). This is also the basis for a directed graph.
    """

    def __init__(self, name: str):
        self.name = name
        if not os.path.isdir(self.name):
            os.mkdir(self.name)
        self.tree = os.listdir(self.name)

    def get_vertices(self):
        """
        returns vertices/nodes of graph
        """
        tree = self.tree
        nodes = []
        for subtree_root in tree:
            root_name = subtree_root.split("orgwiki")[1][:-4]
            with open(f"files/{subtree_root}", "rt") as fh:
                for subtree_branch in fh:
                    if subtree_branch != "\n" and re.match(
                            r"^http://wikipedia\.org//wiki/[\w0-9]+$",
                            subtree_branch):
                        nodes.append(subtree_branch.split("wiki/")[1].strip())
            nodes.append(root_name)
        return nodes

    def edge_list(self):
        """
        returns tuples with edges of the directed graph
        e.g. let G = a graph with nodes A, X, and Y.
        iff A feeds into X and Y, X feeds into Y, and Y feeds into A,
        then edge_list(G) returns [(A, X), (A, Y), (X, Y), (Y, A)]
        """
        tree = self.tree
        nodes = []
        for subtree_root in tree:
            root_name = subtree_root.split("orgwiki")[1][:-4]
            with open(f"files/{subtree_root}", "rt") as fh:
                for subtree_branch in fh:
                    if subtree_branch != "\n" and re.match(r"^http://wikipedia\.org//wiki/[\w0-9]+$", subtree_branch):
                        nodes.append((root_name.strip(), subtree_branch.split("wiki/")[1].strip()))
        return nodes

    def output_degrees(self, vertex):
        """
        from the example above, if our vertex is A.
        then output_degrees(A) will return 2,
        output_degrees(X) will return 1, etc.
        """
        nodes = self.edge_list()
        degrees = 0
        for (i, j) in nodes:
            if i == vertex:
                degrees += 1
        return degrees

    def input_degrees(self, vertex):
        nodes = self.edge_list()
        degrees = 0
        for (i, j) in nodes:
            if j == vertex:
                degrees += 1
        return degrees

    def adjacency_matrix(self):
        wb = xl.load_workbook("thisfile.xlsx")
        ws = wb["Sheet"]
        vertices = self.get_vertices()
        edges = self.edge_list()
        for v in range(len(vertices)):
            ws.cell(row=1, column=v + 2).value = vertices[v]
            ws.cell(row=v + 2, column=1).value = vertices[v]
        for i in range(2, ws.max_row):
            for j in range(2, ws.max_column + 1):
                ws.cell(i, j).value = 0
                edge = (f"{ws.cell(i, 1).value}", f"{ws.cell(1, j).value}")
                if edge in edges:
                    ws.cell(i, j).value = edges.count(edge)
        try:
            wb.save("adjacency_matrix.xlsx")
        except PermissionError:
            pass

        adjacency_matrix = pd.read_excel("adjacency_matrix.xlsx")
        return adjacency_matrix


graph = Graph("files")
graph.adjacency_matrix()
